#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jan  9 11:00:00 2024

@author: Felix Reibold, Steffen Voigtlaender
"""

import gurobipy as gp
import numpy as np
import pandas as pd
import openpyxl
import time

################################################################################################################################################################
# Create output
def Output(m):  
    # Print the result
    status_code = {1:'LOADED', 2:'OPTIMAL', 3:'INFEASIBLE', 4:'INF_OR_UNBD', 5:'UNBOUNDED', 9: 'TIME_LIMIT'} 
    
    status = m.status
    
    print('The optimization status is ' + status_code[status])
    if status == 2 or status == 9:   
        # Retrieve variables value
        print('Optimal solution:')
        for v in m.getVars():
            print(str(v.varName) + " = " + str(v.x))    
        print('Optimal objective value: ' + str(m.objVal) + "\n")
        
# Define MILP Model        
def Model_1_TC (H, muy_m, v_m, co_m, Cap_b, SC, mr_rb, r_a2, r_ia, nk_4, r_a4, m_i, V_i, r_a5, nt_5, l_pF, n_ip, n_pm, demand_i, dWS_b, l_pm, dW_s, dS_sb, station_families, n_f, demand_f, ca, ov, s_ip, wd_p, cl, AL_pb, ntr_2, ht_ip, mv, ALk_pb, q_p, families_parts, L_p, L_s, h_2, w_2, BIG_M_route, mr1_r, mr2_rb, BIG_M_storage, d_p):
    
    start_time = time.time()    
    # Create the optimization model
    model = gp.Model()
    # Set parameters
    model.setParam('OutputFlag', True)
    model.setParam('TimeLimit', 600)

    # Parts
    I = len(assignment_parts_stations)
    # Supermarkets
    B = len(Cap_b)
    #Three vehicle types: forklift, tow train and AGV
    M = 3
    # Two attributes: Size(volume) and weight
    A = 2
    # Possible routes
    R =len(df)
    # Stations
    S = stations
    # Families
    F = int(assignment_parts_stations[-1, 3])
    #Feeding policies (line stocking, boxed supply, sequencing, stationary and travelling kits)
    P = 5
    #Planning horizon
    H = 20
    T = H
    #product demand
    d = H
    #Transportation flows (warehouese - station & supermarket - station)
    W = 2
    W_p = {1: 1, 2: 2, 3: 2, 4: 2, 5: 2}
    
    
    # Define decision variables
    #Binary variable that assigns part i to feeding policy p, vehicle type m and supermarket b
    x_ipmb = model.addVars(I, P, M, B, lb = 0, vtype = gp.GRB.BINARY, name='assignment_parts')
    #Binary variable which indicates if supermarket b is active
    y_b = model.addVars(B, lb = 0, vtype = gp.GRB.BINARY, name='activation_supermarket')
    #Binary variable that decides on the frequency t of vehicle type m when transporting parts assigned to policy p in flow w over route r from supermarket b
    v_pmtrwb = model.addVars(P, M, T, R, W, B, lb = 0, vtype = gp.GRB.BINARY, name='decision_frequency')
    #Binary variable that assigns stations to supermarkets
    q_sb = model.addVars(S, B, lb = 0, vtype = gp.GRB.BINARY, name='assignment_supermarkets')
    #Binary variable that decides if a stationary kit is conveyed by vehicle type m to station s from supermarket b
    k_msb = model.addVars(M, S, B, lb = 0, vtype = gp.GRB.BINARY, name='stationary_kit_vehicle')
    #Binary variable that decides if a travelling kit is conveyed by vehicle type m to the assembly line from supermarket b
    t_mb = model.addVars(M, B, lb = 0, vtype = gp.GRB.BINARY, name='travelling_kit_vehicle')
    #Number of facings for part i assigned to feeding policy p ∈ {1, 2} at station s
    f_ips = model.addVars(I, P, S, lb=0, vtype=gp.GRB.INTEGER, name='number_of_facings_f_ips')
    #Number of facings for part family f assigned to sequencing at station s
    f_fs = model.addVars(F, S, lb = 0, vtype = gp.GRB.INTEGER, name='number_of_facings_f_fs')
    #Number of facings for stationary kitting at station s
    f_s = model.addVars(S, lb = 0, vtype = gp.GRB.INTEGER, name='number_of_facings_f_s')
    #Number of racks for boxed supply in station s
    r_s = model.addVars(S, lb = 0, vtype = gp.GRB.INTEGER, name='number_of_facings_r_s')
    #Costs of providing part i with feeding policy p and vehicle type m to the BoL via supermarket b
    c_ipmb = model.addVars(I, P, M, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='cost_general')
    #Replenishment costs for part i with feeding policy p and vehicle type m via supermarket b
    cR_ipmb = model.addVars(I, P, M, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='replenishment_cost')
    #Preparation costs for part i with feeding policy p and vehicle type m via supermarket b
    cP_ipmb = model.addVars(I, P, M, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='preparation_cost')
    #Transportation costs for part i with feeding policy p and vehicle type m via supermarket b
    cT_ipmb = model.addVars(I, P, M, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='transportation_cost')
    #Usage costs for part i with feeding policy p and vehicle type m via supermarket b
    cU_ipmb = model.addVars(I, P, M, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='usage_cost')
    c_pmtrwb = model.addVars(P, M, T, R, W, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='cost_route')
    #Milk run transportation cost for deliveries made with vehicle type m when transporting parts assigned to policy p in flow w over route r with frequency t
    cT_pmtrwb = model.addVars(P, M, T, R, W, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='transportation_cost_route')
    #Transportation cost for part family f and vehicle type m from supermarket b
    c_fmb = model.addVars(F, M, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='cost_sequencing')
    cT_fmb = model.addVars(F, M, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='transportation_cost_sequencing')
    #Transportation cost for a stationary kit provided by vehicle type m to station s
    c_msb = model.addVars(M, S, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='cost_stationary')
    cT_msb = model.addVars(M, S, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='transportation_cost_stationary')
    #Transportation cost for a travelling kit provided by vehicle type m from supermarket b
    c_mb = model.addVars(M, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='cost_travelling')
    cT_mb = model.addVars(M, B, lb = 0, vtype = gp.GRB.CONTINUOUS, name='transportation_cost_travelling')

    
    
    #Constraints
    #General constarints
    #Each part is assigned to exactly one feeding policy p, one vehicle type m and one supermarket b
    #model.addConstrs( gp.quicksum(x_ipmb[i, p, m, b] for p in range(P) for m in range(M) for b in range(B)) == 1 for i in range(I))
    
    
    #Constraint only active if feeding policy is pre-assigned
    model.addConstrs(gp.quicksum(x_ipmb[i, int(assignment_parts_stations[i, 12] - 1), m, b] for m in range(M) for b in range(B)) == 1 for i in range(I))
    
    
    #boxed supply not possible if parts to big
    model.addConstrs(
        x_ipmb[i, 1, m, b] == 0
        for a in range(A)
        for i in range(I)
        if r_ia[i, a] > r_a2[a]  
        for m in range(M)
        for b in range(B))
    #stationary kits
    model.addConstrs(
        (nk_4 / r_a4[a]) * gp.quicksum(
            ((r_ia[i, a] * m_i[i]) / V_i[i]) * x_ipmb[i, 3, m, b]
            for i, row in enumerate(map(tuple, assignment_parts_stations.tolist()))
            if row[1] == s + 1
        ) <= k_msb[m, s, b]
        for s in range(S)
        for m in range(M)
        for a in range(A)
        for b in range(B))
    #No more than one stationary kit per station and supermarket
    model.addConstrs(
        gp.quicksum(k_msb[m, s, b] for m in range(M))<= q_sb[s, b]
        for b in range(B)
        for s in range(S))
    #travelling kits
    model.addConstrs(
        (nt_5 / r_a5[a]) * gp.quicksum(
            ((r_ia[i, a] * m_i[i]) / V_i[i]) * x_ipmb[i, 4, m, b]
            for i in range(I)
        ) <= t_mb[m, b]
        for s in range(S)
        for m in range(M)
        for a in range(A)
        for b in range(B))
    #No more than one travelling kit for one supermarket
    model.addConstrs(gp.quicksum(t_mb[m, b] for m in range(M)) <= 1 for b in range(B))
    #Parts can only be assigned to a supermarket if supermarket is active
    model.addConstrs(x_ipmb[i, p, m, b] <= y_b[b]  for i in range(I) for b in range(B) for p in range(P) for m in range(M))
    #If a part is assigned to a supermarket, also the station is assigned to the same supermarket
    model.addConstrs((x_ipmb[i, p, m, b] <= q_sb[s, b] for s in range(S) for i, row in enumerate(map(tuple, assignment_parts_stations.tolist())) if row[1] == s + 1 for b in range(B) for p in range(P) for m in range(M)))
    #Stations can only be assigned to a supermarket if the supermarket is active
    model.addConstrs(q_sb[s, b] <= y_b[b]  for s in range(S) for b in range(B))
    #Supermarket have a capacity which refers to the number of parts
    model.addConstrs(gp.quicksum(x_ipmb[i, p, m, b] for i in range(I) for p in range(P) for m in range(M)) <= Cap_b[b] for b in range(B))
    #Parts of the same family are assigned to the same supermarket
    model.addConstrs(
        x_ipmb[i, p, m, b] == x_ipmb[j, p, m, b] 
        for f in range(F)
        for b in range(B)
        for p in range(P)
        for m in range(M)
        for i in range(I) if assignment_parts_stations[i, 3] == f+1
        for j in range(i + 1, I)
        if assignment_parts_stations[i, 3] == assignment_parts_stations[j, 3])
        
    #Route Determination
    #Each station which is assigned to a supermarket is delivered via exactly one route from this supermarket
    model.addConstrs((gp.quicksum(v_pmtrwb[p, m, t, r, w, b] for t in range(T) for r in df[df['Route'].apply(lambda route: s + 1 in route)].index) <= q_sb[s, b] for b in range(B) for s in range(S) for p in range(P) for m in [1, 2] for w in range(W_p[p+1])))
    #Theselection of a suitable route and takt for each vehicle type if any part is assigned to that vehicle type
    model.addConstrs(
            (gp.quicksum(
                x_ipmb[i, p, m, b]
                for i, row in enumerate(map(tuple, assignment_parts_stations.tolist()))
                if row[1] == s + 1
            ) <= BIG_M_route * (
                gp.quicksum(
                    v_pmtrwb[p, m, t, r, w, b]
                    for t in range(T)
                    for r in df[df['Route'].apply(lambda route: s + 1 in route)].index))
            for p in range(P)
            for m in [1, 2]
            for s in range(S)
            for b in range(B)
            for w in range(W_p[p+1])))
    
    #Line-sided storage
    #ensures that containers used for line stocked and boxed supply parts, delivered by vehicle type m ∈ M′ to each station, between consecutive deliveries, do not exceed the number of facings reserved for them
    model.addConstrs(
            (gp.quicksum(
                ((((demand_i[i] * t) / (H * n_ip[i, p] * d_p[p])) * x_ipmb[i, p, m, b])) - BIG_M_storage + (BIG_M_storage * (gp.quicksum(v_pmtrwb[p, m, t, r, w, b] for r in df[df['Route'].apply(lambda route: s + 1 in route)].index)))
                for b in range(B)
            ) <= f_ips[i, p, s])
        for p in [0, 1]
        for s in range(S)
        for i, row in enumerate(map(tuple, assignment_parts_stations.tolist())) if row[1] == s + 1
        for t in range(T)
        for m in [1, 2]
        for w in range(W_p[p+1]))
    #determines the number of racks used for boxed supply
    model.addConstrs(
            ((1 / (h_2 * w_2)) * gp.quicksum(f_ips[i, 1, s] for i, row in enumerate(map(tuple, assignment_parts_stations.tolist())) if row[1] == s + 1)
            <= r_s[s]
            for s in range(S)))
    #number of containers used for sequencing and stationary kitting delivered by vehicle type m ∈ M′ to each station, between consecutive deliveries do not exceed their respective facings
    model.addConstrs(
            (gp.quicksum(
                ((((demand_f[f] * t) / (H * n_f[f])) * x_ipmb[i, 2, m, b])) - BIG_M_storage + (BIG_M_storage * (gp.quicksum(v_pmtrwb[2, m, t, r, 1, b] for r in df[df['Route'].apply(lambda route: s + 1 in route)].index)))
                for b in range(B)
            ) <= f_fs[f, s])
        for s in range(S)
        for f in station_families[s+1]
        for i in families_parts[f+1] if i > 0 and (i == 1 or assignment_parts_stations[i-1,3] != assignment_parts_stations[i,3])
        for t in range(T)
        for m in [1, 2])
    model.addConstrs(
            (gp.quicksum(
                (((t / nk_4) * k_msb[m, s, b]) - BIG_M_storage + (BIG_M_storage * gp.quicksum(v_pmtrwb[3, m, t, r, 1, b] for r in df[df['Route'].apply(lambda route: s + 1 in route)].index)))
                for b in range(B)
            ) <= f_s[s])
        for s in range(S)
        for t in range(T)
        for m in [1, 2])
    #lower bound on the number of facings available
    model.addConstrs(
                (q_p[p] * gp.quicksum(x_ipmb[i, p, m, b] for m in range(M) for b in range(B)) <= f_ips[i, p, s]
                for p in [0, 1]
                for s in range(S)
                for i, row in enumerate(map(tuple, assignment_parts_stations.tolist())) if row[1] == s + 1))
    #lower bound on the number of facings available
    model.addConstrs(
            (q_p[2] * gp.quicksum(x_ipmb[i, 2, m, b] for m in range(M) for b in range(B)) <= f_fs[f, s]
            for s in range(S)
            for f in station_families[s+1]
            for i in families_parts[f+1] if i > 0 and (i == 1 or assignment_parts_stations[i-1,3] != assignment_parts_stations[i,3])))
    #lower bound on the number of facings available
    model.addConstrs(
            (q_p[3] * gp.quicksum(k_msb[m, s, b] for m in range(M) for b in range(B)) <= f_s[s]
            for s in range(S)))
    #ensures that the summed length of all facings do not violate the space available at any station
    model.addConstrs(
            (gp.quicksum(L_p[0] * f_ips[i, 0, s] for i, row in enumerate(map(tuple, assignment_parts_stations.tolist())) if row[1] == s + 1)
            + L_p[1] * r_s[s]
            + gp.quicksum(L_p[2] * f_fs[f, s] for f in station_families[s+1])
            + L_p[3] * f_s[s]
            <= L_s
            for s in range(S)))  

    #Cost Parameter caluclation - equations
    #Costs of providing part i with feeding policy p and vehicle type m to the BoL via supermarket b
    model.addConstrs(c_ipmb[i, p, m, b] == (cR_ipmb[i, p, 0, b] + cP_ipmb[i, p, m, b] + cT_ipmb[i, p, 0, b] + cU_ipmb[i, p, m, b]) for i in range(I) for p in range(P) for m in range(M) for b in range(B))
    #costs incurred for each part transported by AGVs or tow trains
    model.addConstrs(c_pmtrwb[p, m, t, r, w, b] == cT_pmtrwb[p, m, t, r, w, b] for p in range(P) for m in [1, 2] for t in range(T) for r in range(R) for b in range(B) for w in range(W))
    #transportation cost for sequenced parts
    model.addConstrs(c_fmb[f, m, b] == cT_fmb[f, m, b] for f in range(F) for m in range(M) for b in range(B))
    #transportation costs for stationary kits
    model.addConstrs(c_msb[m, s, b] == cT_msb[m, s, b] for m in range(M) for s in range(S) for b in range(B))
    #transportation costs for travelling kits
    model.addConstrs(c_mb[m, b] == cT_mb[m, b] for m in range(M) for b in range(B))
    
    #Replenishment costs
    model.addConstrs(
            (cR_ipmb[i, p, 0, b] == (co_m[0] * ((demand_i[i] / n_ip[i, 0]) * ((dWS_b[b] / (n_pm[0, 0] * muy_m[0] * v_m[0])) + l_pF[p])))
            for i in range(I)
            for p in [1, 2, 3, 4]
            for b in range(B)))
    
    #Preparation costs
    model.addConstrs(
        cP_ipmb[i, 1, m, b] == (cl * ((demand_i[i] / n_ip[i, 0]) * (AL_pb[1] / (ntr_2 * ov)) + (ht_ip + s_ip) * demand_i[i]))
        for i in range(I)
        for m in range(M)
        for b in range(B))
    model.addConstrs(
        cP_ipmb[i, 2, m, b] == (cl * ((demand_i[i] / n_ip[i, 1]) * ((AL_pb[2] + mv) / ov) + (ht_ip + s_ip) * demand_i[i]))
        for i in range(I)
        for m in range(M)
        for b in range(B))
    model.addConstrs(cP_ipmb[i, 3, m, b] == (cl * (((ALk_pb[3] * d)/ (ov * nk_4)) + (ht_ip + s_ip) * demand_i[i]))for i in range(I) for m in range(M) for b in range(B))
    model.addConstrs(cP_ipmb[i, 4, m, b] == (cl * (((ALk_pb[4] * d)/ (ov * nt_5)) + (ht_ip + s_ip) * demand_i[i]))for i in range(I) for m in range(M) for b in range(B))

    #Transportation costs
    model.addConstrs(
            (cT_ipmb[i, 0, m, b] == co_m[m] * ((demand_i[i] / n_ip[i, 0]) * ((dW_s[s] / (n_pm[0, m] * muy_m[m] * v_m[m])) + l_pm[0, m]))
            for s in range(S)
            for i, row in enumerate(map(tuple, assignment_parts_stations.tolist())) if row[1] == s + 1
            for m in range(M)
            for b in range(B)))
    model.addConstrs(
            (cT_ipmb[i, 1, m, b] == co_m[m] * ((demand_i[i] / n_ip[i, 1]) * ((dS_sb[b, s] / (n_pm[1, m] * muy_m[m] * v_m[m])) + l_pm[1, m]))
            for s in range(S)
            for i, row in enumerate(map(tuple, assignment_parts_stations.tolist())) 
            if row[1] == s + 1
            for m in range(M)
            for b in range(B)))
    model.addConstrs(
            (cT_fmb[f, m, b] == co_m[m] * ((demand_f[f] / n_f[f]) * ((dS_sb[b, s] / (n_pm[2, m] * muy_m[m] * v_m[m])) + l_pm[2, m]))
            for s in range(S)
            for f in station_families[s+1]
            for m in range(M)
            for b in range(B)))
    model.addConstrs(
            (cT_msb[m, s, b] == co_m[m] * ((d / nk_4) * ((dS_sb[b, s] / (n_pm[3, m] * muy_m[m] * v_m[m])) + l_pm[3, m]))
            for s in range(S)
            for m in range(M)
            for b in range(B)))
    model.addConstrs(
            (cT_mb[m, b] == co_m[m] * ((d / nt_5) * ((dS_sb[b, 1] / (n_pm[4, m] * muy_m[m] * v_m[m])) + l_pm[4, m]))
            for m in range(M)
            for b in range(B)))
    model.addConstrs(
            (cT_pmtrwb[0, m, t, r, 0, 0] == (co_m[m] * ((H * mr1_r[r]) / ((t+1) * muy_m[m] * v_m[m])))
            for m in [1, 2]
            for t in range(T)
            for r in range(R)))
    model.addConstrs(
            (cT_pmtrwb[p, m, t, r, 1, b] == co_m[m] * ((H * mr2_rb[r, b]) / ((t+1) * muy_m[m] * v_m[m]))
            for p in [1, 2, 3, 4]
            for m in [1, 2]
            for t in range(T)
            for r in range(R)
            for b in range(B)))

    #Usage costs
    model.addConstrs(cU_ipmb[i, p, m, b] == (ca * demand_i[i] * ((wd_p/ov) + s_ip)) for i in range(I) for p in [0, 1] for m in range(M) for b in range(B))
    model.addConstrs(cU_ipmb[i, p, m, b] == (ca * demand_i[i] * (wd_p/ov)) for i in range(I) for p in [2, 3] for m in range(M) for b in range(B))
    
    
    # Define objective function
    # Part 1: cost of feeding part i with feeding policy p through vehicle type m, denoted by cipmb.
    # This cost covers all processes, namely replenishment, preparation, transportation, and usage cost.
    # However, this transportation cost only includes transportation cost for line stocked and boxed supply parts delivered by forklift
    transportation_costs_forklift = gp.quicksum(c_ipmb[i, p, m, b] * x_ipmb[i, p, m, b] for i in range(I) for p in range(P) for m in range(M) for b in range(B))
    #Part 2: cost of transporting all parts assigned to an AGV or a tow train between the warehouse or the supermarket and the line, denoted by cpmtrwb
    transportation_costs_TT_and_AGV = gp.quicksum(c_pmtrwb[p, m, t, r, w, b] * v_pmtrwb[p, m, t, r, w, b] for p in range(P) for m in [1, 2] for t in range(T) for r in range(R) for w in range(W) for b in range(B))
    #Part 3: cost of transporting sequenced parts by forklift cfm
    transportation_costs_sequenced_parts = gp.quicksum(c_fmb[f, m, b] * x_ipmb[i, 2, m, b] for f in range(F) for i in families_parts.get(f, []) if i > 0 and (i == 1 or assignment_parts_stations[i-1,3] != assignment_parts_stations[i,3]) for m in range(M) for b in range(B))
    #Part 4: transportation cost of kitted parts by a forklift for stationary kits cmsb
    transportation_costs_stationary_kits = gp.quicksum(c_msb[m, s, b] * k_msb[m, s, b] for m in range(M) for s in range(S) for b in range(B))
    #Part 5: transportation cost of kitted parts by a forklift for travelling kits cmb
    transportation_costs_travelling_kits = gp.quicksum(c_mb[m, b] * t_mb[m, b] for m in range(M) for b in range(B))
    #Part 6: supermarket costs: activation and capacity related
    fix_cost_supermarket = gp.quicksum(((SC * Cap_b[b]) + Fix_S) * y_b[b] for b in range(B))

    model.setObjective(transportation_costs_forklift   + transportation_costs_TT_and_AGV + transportation_costs_sequenced_parts + transportation_costs_stationary_kits + transportation_costs_travelling_kits + fix_cost_supermarket, sense = gp.GRB.MINIMIZE)
    obj_value = model.getObjective().getValue()
    
    # Optimize the model
    model.optimize()
    #model.write('mp.lp')
    model.write('MILP_Experiment_41.sol')
    Output(model)
    # Record end time
    end_time = time.time()
    # Calculate and print the duration
    duration = end_time - start_time
    print("Optimization duration: {:.2f} seconds".format(duration))
    
    return obj_value
    

# Read Input and data preparation

excel_file_path = 'Input_Data.xlsx'
Input = openpyxl.load_workbook(excel_file_path)
sheet = Input['Tabelle1']
#Utilisation rate of vehicle type m
muy_m = np.array([cell.value for cell in sheet['B4:D4'][0]])
#Velocity of vehicle type m
v_m = np.array([cell.value for cell in sheet['B5:D5'][0]])
#Planning horizon
H = sheet['B6'].value
#Cost of moving vehicle type m on the shop floor
co_m = np.array([cell.value for cell in sheet['B7:D7'][0]])
#Wage costs for supermarket associate
Fix_S = sheet['B32'].value
#Capacity of supermarket b
Cap_b = [cell.value for cell in sheet['B8:C8'][0]]
#costs per part capacity unit in supermarket (rent)
SC = sheet['B9'].value
#Number of stationary kits that fit into a kit container
nk_4 = sheet['B12'].value
#Number of travelling kits that fit into a kit container
nt_5 = sheet['B14'].value
#Number of shelves in a rack used for boxed supply containers (number of containers)
h_2 = sheet['B16'].value
#Number of boxes fitting in a rack shelf used for boxed supply
w_2 = sheet['B17'].value
#Minimum number of facings for containers of feeding policy p,
q_p = np.array([cell.value for cell in sheet['B18:F18'][0]])
#Length of the BoL at station s
L_s = sheet['B19'].value 
#Length of facing for feeding policy p ∈ {1, 2, 3, 4}
L_p = np.array([cell.value for cell in sheet['B20:F20'][0]]) 
#Wage of a logistical operator
cl = sheet['B22'].value 
#Aisle length in the preparation area corresponding to line feeding policy p
AL_pb = np.array([cell.value for cell in sheet['B23:F23'][0]])  
#Batch size: Number of bins that fit into a trolley in the preparation area   
ntr_2 = sheet['B24'].value 
#Walking velocity of an operator
ov = sheet['B25'].value 
#Walking distance in the preparation area between several variants of the same family
mv =  sheet['B28'].value 
#Aisle length in the preparation area corresponding to kitted parts
ALk_pb = np.array([cell.value for cell in sheet['B29:F29'][0]]) 
#Walking distance at the assembly line to pick parts corresponding to feeding policy p ∈ {1, 2, 3, 4}
wd_p =  sheet['B30'].value 
#Wage of an assembly operator
ca = sheet['B31'].value
#BIG M values
BIG_M_route = sheet['B33'].value  
BIG_M_storage = sheet['B34'].value 
#Search time for part i corresponding to line feeding policy p at the preparation area and BoL
s_ip = sheet['B37'].value 
#Preparation handling time required for part i corresponding to line feeding policy p ∈ P′
ht_ip = sheet['B38'].value 
#Forklift loading and unloading time for feeding policy p
l_pF = np.array([cell.value for cell in sheet['B21:F21'][0]]) 
#Depth of rack used for feeding policy p (number of containers)
d_p = np.array([cell.value for cell in sheet['B35:C35'][0]])  
#Two-way distance between warehouse and supermarket
dWS_b = np.array([cell.value for cell in sheet['B36:C36'][0]]) 

excel_file_path_distance_matrix = 'distance_matrix_10x2_storage_to_stations.xlsx'
distance_matrix = pd.read_excel(excel_file_path_distance_matrix).to_numpy()
distance_matrix_1 = distance_matrix[1:3, :]
station_length = sheet['B10'].value
stations = distance_matrix_1.shape[1]

#Maximum volume possible in a container of policy p and Maximum weight allowed in a container of policy p
r_a2 = np.array([cell.value for cell in sheet['B11:C11'][0]]).astype(float)
r_a4 = np.array([cell.value for cell in sheet['B13:C13'][0]]).astype(float)
r_a5 = np.array([cell.value for cell in sheet['B15:C15'][0]]).astype(float)


milk_run_lengths = []

#Route creation and definition
for i in range(stations):
    for j in range(i + 1, stations + 1):
        route = list(range(i + 1, j + 1))
        length = [0] * len(Cap_b)  
        for b in range(len(Cap_b)):
            length[b] = distance_matrix_1[b][route[0] - 1] 
            for k in range(len(route) - 1):
                length[b] += station_length 
            length[b] += distance_matrix_1[b][route[-1] - 1] 
        milk_run_lengths.append((route, length))

routes_array = np.array(milk_run_lengths, dtype=object)
split_lengths = [(route, length[0], length[1]) for route, length in routes_array]
df = pd.DataFrame(split_lengths, columns=['Route', 'Length_1', 'Length_2'])
dW_s = distance_matrix[0, :] * 2 
dS_sb = distance_matrix[1:3, :] * 2             
excel_file_path_assignment_parts_stations = 'part_station_assignment_10_02.xlsx'
assignment_parts_stations = pd.read_excel(excel_file_path_assignment_parts_stations)
assignment_parts_stations = assignment_parts_stations.to_numpy()    
columns_5_6 = assignment_parts_stations[:, 5:7]
r_ia = np.array(columns_5_6)   
m_i = np.array(assignment_parts_stations[:, 7])
V_i = np.array(assignment_parts_stations[:, 8])
n_ip = np.array(assignment_parts_stations[:, 10:12])   
demand_i = np.array(assignment_parts_stations[:, 9])                                                
mr_rb = np.array(df[[f'Length_{i+1}' for i in range(len(Cap_b))]])

#Number of bins or containers that fit into a vehicle type m
excel_file_path_n_pm = 'n_pm.xlsx'
Input_n_pm = openpyxl.load_workbook(excel_file_path_n_pm)
Input_n_pm = Input_n_pm['Tabelle1']
num_rows = 5
num_cols = 3
n_pm = [[Input_n_pm.cell(row=i, column=j).value for j in range(1, num_cols + 1)] for i in range(1, num_rows + 1)]
n_pm = np.array(n_pm)

#l_pm
excel_file_path_l_pm = 'l_pm.xlsx'
Input_l_pm = openpyxl.load_workbook(excel_file_path_l_pm)
Input_l_pm = Input_l_pm['Tabelle1']
l_pm = np.array([[Input_l_pm.cell(row=i, column=j).value for j in range(1, num_cols + 1)] for i in range(1, num_rows + 1)])

station_families = {}
for row in assignment_parts_stations:
    station = int(row[1])  
    family = int(row[3])  
    family -= 1
    if station not in station_families:
        station_families[station] = set()
    station_families[station].add(family)
    
families_parts = {}
for row in assignment_parts_stations:
    part = int(row[0])  
    family = int(row[3])  
    part -= 1
    if family not in families_parts:
        families_parts[family] = set()
    families_parts[family].add(part)
    
#demand_f
excel_file_path = 'n_f.xlsx'
Input_nf = openpyxl.load_workbook(excel_file_path)
Input_nf = Input_nf['Tabelle1']
last_row = Input_nf.max_row
while Input_nf.cell(row=last_row, column=1).value is None and last_row > 1:
    last_row -= 1
values_a = [Input_nf.cell(row=i, column=1).value for i in range(2, last_row + 1)]
values_b = [Input_nf.cell(row=i, column=2).value for i in range(2, last_row + 1)]

#family demand
demand_f = np.array(values_a)
n_f = np.array(values_b)


#Milk run length for route r from the warehouse
milk_run_lengths_w1 = []
distance_matrix_1 = distance_matrix[0, :]
for i in range(stations):
    for j in range(i + 1, stations + 1):
        route_w1 = list(range(i + 1, j + 1))
        route_length = [0] * 1  
        route_length = distance_matrix_1[route_w1[0] - 1] 
        for k in range(len(route_w1) - 1):
                route_length += station_length  
        route_length += distance_matrix_1[route_w1[-1] - 1]  
        milk_run_lengths_w1.append((route_w1, route_length))
routes_array_w1 = np.array(milk_run_lengths_w1, dtype=object)
df_w1 = pd.DataFrame(routes_array_w1, columns=['Route', 'Length'])
mr1_r = df_w1['Length'].to_numpy(dtype=float)

#Milk run length for route r from the supermarket 
mr2_rb = df[['Length_1', 'Length_2']]
mr2_rb = mr2_rb.values

OptModel_low = Model_1_TC(H, muy_m, v_m, co_m, Cap_b, SC, mr_rb, r_a2, r_ia, nk_4, r_a4, m_i, V_i, r_a5, nt_5, l_pF, n_ip, n_pm, demand_i, dWS_b, l_pm, dW_s, dS_sb, station_families, n_f, demand_f, ca, ov, s_ip, wd_p, cl, AL_pb, ntr_2, ht_ip, mv, ALk_pb, q_p, families_parts, L_p, L_s, h_2, w_2, BIG_M_route, mr1_r, mr2_rb, BIG_M_storage, d_p)    
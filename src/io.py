#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jan  9 11:00:00 2024

@author: Felix Reibold, Steffen Voigtlaender
"""

from __future__ import annotations

from pathlib import Path
import numpy as np
import pandas as pd
import openpyxl





def _read_row(sheet, excel_range: str) -> np.ndarray:
    return np.array([cell.value for cell in sheet[excel_range][0]])


def _read_scalar(sheet, cell_ref: str):
    return sheet[cell_ref].value


def _read_matrix_xlsx(path: Path) -> np.ndarray:
    return pd.read_excel(path).to_numpy()




def _build_routes(distance_matrix_1: np.ndarray, station_length: float, stations: int, Cap_b: list[float]) -> pd.DataFrame:
    milk_run_lengths = []

    #Route creation and definition
    for i in range(stations):
        for j in range(i + 1, stations + 1):
            route = list(range(i + 1, j + 1))
            length = [0] * len(Cap_b)
            for b in range(len(Cap_b)):
                length[b] = distance_matrix_1[b][route[0] - 1]
                for k in range(len(route) - 1):
                    length[b] += station_length
                length[b] += distance_matrix_1[b][route[-1] - 1]
            milk_run_lengths.append((route, length))

    # build dataframe with Length_1..Length_B dynamically
    cols = ["Route"] + [f"Length_{b+1}" for b in range(len(Cap_b))]
    rows = []
    for route, length in milk_run_lengths:
        rows.append([route] + list(length))

    return pd.DataFrame(rows, columns=cols)


def _build_routes_warehouse(distance_matrix_row: np.ndarray, station_length: float, stations: int) -> np.ndarray:
    #Milk run length for route r from the warehouse
    milk_run_lengths_w1 = []
    for i in range(stations):
        for j in range(i + 1, stations + 1):
            route_w1 = list(range(i + 1, j + 1))
            route_length = distance_matrix_row[route_w1[0] - 1]
            for k in range(len(route_w1) - 1):
                route_length += station_length
            route_length += distance_matrix_row[route_w1[-1] - 1]
            milk_run_lengths_w1.append((route_w1, route_length))

    df_w1 = pd.DataFrame(milk_run_lengths_w1, columns=["Route", "Length"])
    return df_w1["Length"].to_numpy(dtype=float)


def _build_station_families(assignment_parts_stations: np.ndarray) -> dict[int, set[int]]:
    station_families: dict[int, set[int]] = {}
    for row in assignment_parts_stations:
        station = int(row[1])
        family = int(row[3])
        family -= 1
        if station not in station_families:
            station_families[station] = set()
        station_families[station].add(family)
    return station_families


def _build_families_parts(assignment_parts_stations: np.ndarray) -> dict[int, set[int]]:
    families_parts: dict[int, set[int]] = {}
    for row in assignment_parts_stations:
        part = int(row[0])
        family = int(row[3])
        part -= 1
        if family not in families_parts:
            families_parts[family] = set()
        families_parts[family].add(part)
    return families_parts


def prepare_inputs(
    input_data_xlsx: Path,
    n_pm_xlsx: Path,
    l_pm_xlsx: Path,
    n_f_xlsx: Path,
    distance_matrix_xlsx: Path,
    part_station_assignment_xlsx: Path,
):
    # Read Input and data preparation

    Input = openpyxl.load_workbook(input_data_xlsx)
    sheet = Input["Tabelle1"]

    distance_matrix = _read_matrix_xlsx(distance_matrix_xlsx)
    distance_matrix_1 = distance_matrix[1:1 + (distance_matrix.shape[0] - 1), :]

    #Utilisation rate of vehicle type m
    muy_m = _read_row(sheet, "B4:D4")
    #Velocity of vehicle type m
    v_m = _read_row(sheet, "B5:D5")
    #Planning horizon
    H = _read_scalar(sheet, "B6")
    #Cost of moving vehicle type m on the shop floor
    co_m = _read_row(sheet, "B7:D7")
    #Wage costs for supermarket associate
    Fix_S = _read_scalar(sheet, "B32")
    #Capacity of supermarket b
    Cap_b = [cell.value for cell in sheet["B8:C8"][0]][: distance_matrix_1.shape[0]]
    #costs per part capacity unit in supermarket (rent)
    SC = _read_scalar(sheet, "B9")
    #Number of stationary kits that fit into a kit container
    nk_4 = _read_scalar(sheet, "B12")
    #Number of travelling kits that fit into a kit container
    nt_5 = _read_scalar(sheet, "B14")
    #Number of shelves in a rack used for boxed supply containers (number of containers)
    h_2 = _read_scalar(sheet, "B16")
    #Number of boxes fitting in a rack shelf used for boxed supply
    w_2 = _read_scalar(sheet, "B17")
    #Minimum number of facings for containers of feeding policy p,
    q_p = _read_row(sheet, "B18:F18")
    #Length of the BoL at station s
    L_s = _read_scalar(sheet, "B19")
    #Length of facing for feeding policy p ∈ {1, 2, 3, 4}
    L_p = _read_row(sheet, "B20:F20")
    #Wage of a logistical operator
    cl = _read_scalar(sheet, "B22")
    #Aisle length in the preparation area corresponding to line feeding policy p
    AL_pb = _read_row(sheet, "B23:F23")
    #Batch size: Number of bins that fit into a trolley in the preparation area
    ntr_2 = _read_scalar(sheet, "B24")
    #Walking velocity of an operator
    ov = _read_scalar(sheet, "B25")
    #Walking distance in the preparation area between several variants of the same family
    mv = _read_scalar(sheet, "B28")
    #Aisle length in the preparation area corresponding to kitted parts
    ALk_pb = _read_row(sheet, "B29:F29")
    #Walking distance at the assembly line to pick parts corresponding to feeding policy p ∈ {1, 2, 3, 4}
    wd_p = _read_scalar(sheet, "B30")
    #Wage of an assembly operator
    ca = _read_scalar(sheet, "B31")
    #BIG M values
    BIG_M_route = _read_scalar(sheet, "B33")
    BIG_M_storage = _read_scalar(sheet, "B34")
    #Search time for part i corresponding to line feeding policy p at the preparation area and BoL
    s_ip = _read_scalar(sheet, "B37")
    #Preparation handling time required for part i corresponding to line feeding policy p ∈ P′
    ht_ip = _read_scalar(sheet, "B38")
    #Forklift loading and unloading time for feeding policy p
    l_pF = _read_row(sheet, "B21:F21")
    #Depth of rack used for feeding policy p (number of containers)
    d_p = np.array([cell.value for cell in sheet["B35:C35"][0]])
    #Two-way distance between warehouse and supermarket
    dWS_b = np.array([cell.value for cell in sheet["B36:C36"][0]])

    station_length = _read_scalar(sheet, "B10")
    stations = distance_matrix_1.shape[1]

    #Maximum volume possible in a container of policy p and Maximum weight allowed in a container of policy p
    r_a2 = np.array([cell.value for cell in sheet["B11:C11"][0]]).astype(float)
    r_a4 = np.array([cell.value for cell in sheet["B13:C13"][0]]).astype(float)
    r_a5 = np.array([cell.value for cell in sheet["B15:C15"][0]]).astype(float)

    df = _build_routes(distance_matrix_1, station_length, stations, Cap_b)
    dW_s = distance_matrix[0, :] * 2
    dS_sb = distance_matrix[1:1 + distance_matrix_1.shape[0], :] * 2



    assignment_parts_stations = pd.read_excel(part_station_assignment_xlsx).to_numpy()
    columns_5_6 = assignment_parts_stations[:, 5:7]
    r_ia = np.array(columns_5_6)
    m_i = np.array(assignment_parts_stations[:, 7])
    V_i = np.array(assignment_parts_stations[:, 8])
    n_ip = np.array(assignment_parts_stations[:, 10:12])
    demand_i = np.array(assignment_parts_stations[:, 9])
    mr_rb = np.array(df[[f"Length_{i+1}" for i in range(len(Cap_b))]])

    #Number of bins or containers that fit into a vehicle type m
    Input_n_pm = openpyxl.load_workbook(n_pm_xlsx)["Tabelle1"]
    num_rows = 5
    num_cols = 3
    n_pm = [[Input_n_pm.cell(row=i, column=j).value for j in range(1, num_cols + 1)] for i in range(1, num_rows + 1)]
    n_pm = np.array(n_pm)

    #l_pm
    Input_l_pm = openpyxl.load_workbook(l_pm_xlsx)["Tabelle1"]
    l_pm = np.array([[Input_l_pm.cell(row=i, column=j).value for j in range(1, num_cols + 1)] for i in range(1, num_rows + 1)])

    station_families = _build_station_families(assignment_parts_stations)
    families_parts = _build_families_parts(assignment_parts_stations)

    #demand_f
    Input_nf = openpyxl.load_workbook(n_f_xlsx)["Tabelle1"]
    last_row = Input_nf.max_row
    while Input_nf.cell(row=last_row, column=1).value is None and last_row > 1:
        last_row -= 1
    values_a = [Input_nf.cell(row=i, column=1).value for i in range(2, last_row + 1)]
    values_b = [Input_nf.cell(row=i, column=2).value for i in range(2, last_row + 1)]

    #family demand
    demand_f = np.array(values_a)
    n_f = np.array(values_b)

    mr1_r = _build_routes_warehouse(distance_matrix[0, :], station_length, stations)

    #Milk run length for route r from the supermarket
    mr2_rb = df[[f"Length_{i+1}" for i in range(len(Cap_b))]].values

    return {
        "sheet": sheet,
        "muy_m": muy_m,
        "v_m": v_m,
        "H": H,
        "co_m": co_m,
        "Fix_S": Fix_S,
        "Cap_b": Cap_b,
        "SC": SC,
        "nk_4": nk_4,
        "nt_5": nt_5,
        "h_2": h_2,
        "w_2": w_2,
        "q_p": q_p,
        "L_s": L_s,
        "L_p": L_p,
        "cl": cl,
        "AL_pb": AL_pb,
        "ntr_2": ntr_2,
        "ov": ov,
        "mv": mv,
        "ALk_pb": ALk_pb,
        "wd_p": wd_p,
        "ca": ca,
        "BIG_M_route": BIG_M_route,
        "BIG_M_storage": BIG_M_storage,
        "s_ip": s_ip,
        "ht_ip": ht_ip,
        "l_pF": l_pF,
        "d_p": d_p,
        "dWS_b": dWS_b,
        "distance_matrix": distance_matrix,
        "station_length": station_length,
        "stations": stations,
        "r_a2": r_a2,
        "r_a4": r_a4,
        "r_a5": r_a5,
        "df": df,
        "dW_s": dW_s,
        "dS_sb": dS_sb,
        "assignment_parts_stations": assignment_parts_stations,
        "r_ia": r_ia,
        "m_i": m_i,
        "V_i": V_i,
        "n_ip": n_ip,
        "demand_i": demand_i,
        "mr_rb": mr_rb,
        "n_pm": n_pm,
        "l_pm": l_pm,
        "station_families": station_families,
        "families_parts": families_parts,
        "demand_f": demand_f,
        "n_f": n_f,
        "mr1_r": mr1_r,
        "mr2_rb": mr2_rb,
    }


